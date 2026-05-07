"""
Build pre-fill validation Excel workbook (8 sheets) before DOCX fill.

Date cells use MM/DD/YYYY where a column is date-oriented.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from shared_core.extractor import FieldCandidate
from shared_core.mapping import DIRECT_VALUE_FIELDS, FIELD_ALIASES

from .date_format import normalize_date_mmddyyyy
from .register_heuristics import (
    harvest_employee_function_training,
    harvest_forms,
    harvest_measure_and_test_equipment,
    harvest_ndt_technician_qualifications,
    harvest_procedures,
)
from .workbook_layout import B2_SECTION_CODE, sheet_category_for_field


# B-2 direct fields whose extracted values are normalized as dates in the workbook.
_DIRECT_DATE_FIELDS = {
    "permission_date",
    "date_approved",
    "training_date",
    "calibration_date",
    "calibration_due",
    "date_qualified",
    "qualification_expiration_date",
    "visual_acuity_date",
    "visual_acuity_due",
}


def _norm_value_for_field(field: str, val: str) -> str:
    if field in _DIRECT_DATE_FIELDS:
        return normalize_date_mmddyyyy(val)
    return val


def _autosize(ws, max_width: int = 56) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max_width, max(10, length + 2))


def _write_sheet(
    ws,
    headers: list[str],
    rows: list[dict[str, Any]],
    *,
    date_columns: frozenset[str] | None = None,
) -> None:
    date_columns = date_columns or frozenset()
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for ri, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            val = row.get(h, "")
            if h in date_columns:
                val = normalize_date_mmddyyyy(val)
            ws.cell(row=ri, column=c, value=val)


def _preview_label(field: str) -> str:
    aliases = FIELD_ALIASES.get(field, [field])
    return aliases[0] if aliases else field


def _src_name(path: str) -> str:
    try:
        return Path(path).name
    except Exception:
        return path or ""


def build_validation_workbook(
    *,
    candidates: list[FieldCandidate],
    blocks: list[dict],
    source_path: Path,
    form_code: str,
    output_path: Path,
    merged_extracted: dict[str, Any] | None = None,
    ml_info: dict[str, Any] | None = None,
) -> Path:
    merged_extracted = dict(merged_extracted or {})
    ml_info = dict(ml_info or {})
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    source_name = source_path.name

    wb = Workbook()
    wb.remove(wb.active)

    proc_rows = harvest_procedures(blocks, str(source_path))
    form_rows = harvest_forms(blocks, str(source_path))
    mte_rows = harvest_measure_and_test_equipment(blocks, str(source_path))
    emp_rows = harvest_employee_function_training(blocks, str(source_path))
    ndt_rows = harvest_ndt_technician_qualifications(blocks, str(source_path))

    _write_sheet(
        wb.create_sheet("PROCEDURES"),
        ["procedure_name", "procedure_id", "approver", "date_approved", "rev"],
        proc_rows,
        date_columns=frozenset({"date_approved"}),
    )
    _write_sheet(
        wb.create_sheet("FORMS"),
        ["form_name", "form_id", "approver", "date_approved", "rev"],
        form_rows,
        date_columns=frozenset({"date_approved"}),
    )
    _write_sheet(
        wb.create_sheet("MEASURE_AND_TEST_EQUIPMENT"),
        [
            "measure_and_test_equipment",
            "id",
            "function_performed",
            "calibration_date",
            "calibration_due",
        ],
        mte_rows,
        date_columns=frozenset({"calibration_date", "calibration_due"}),
    )
    _write_sheet(
        wb.create_sheet("EMPLOYEE_FUNCTION_TRAINING"),
        ["employee_name", "function_performed", "date_received", "function_training_date"],
        emp_rows,
        date_columns=frozenset({"date_received", "function_training_date"}),
    )
    _write_sheet(
        wb.create_sheet("NDT_TECHNICIAN_QUALIFICATIONS"),
        [
            "ndt_technician_id",
            "level_qualified",
            "ndt_methods",
            "date_qualified",
            "qualification_expiration_date",
            "date_of_visual_acuity_exam",
            "due_date_visual_acuity_exam",
        ],
        ndt_rows,
        date_columns=frozenset(
            {
                "date_qualified",
                "qualification_expiration_date",
                "date_of_visual_acuity_exam",
                "due_date_visual_acuity_exam",
            }
        ),
    )

    by_field: dict[str, list[FieldCandidate]] = defaultdict(list)
    for c in candidates:
        if c.b2_field in DIRECT_VALUE_FIELDS:
            by_field[c.b2_field].append(c)

    b2_rows: list[dict[str, Any]] = []
    for field in sorted(DIRECT_VALUE_FIELDS):
        label = _preview_label(field)
        cat = sheet_category_for_field(field)
        clist = by_field.get(field, [])
        distinct_vals: list[str] = []
        seen: set[str] = set()
        for c in clist:
            v = (c.value or "").strip()
            if v and v not in seen:
                seen.add(v)
                distinct_vals.append(v)

        ml_note = ""
        if ml_info.get("ml_engine") == "semantic_tfidf" and field in merged_extracted and merged_extracted.get(field):
            ml_note = "post_merge_includes_ml_semantic"

        if not clist:
            ev = merged_extracted.get(field, "") or ""
            ev = _norm_value_for_field(field, str(ev))
            b2_rows.append(
                {
                    "b2_field": label,
                    "extracted_value": ev,
                    "source_document": _src_name(source_name) if ev else "",
                    "source_page": "",
                    "sheet_category": cat,
                    "match_status": "missing" if not ev else "matched",
                    "notes": ml_note,
                }
            )
            continue

        if len(distinct_vals) <= 1:
            c0 = clist[0]
            val0 = _norm_value_for_field(field, distinct_vals[0] if distinct_vals else "")
            b2_rows.append(
                {
                    "b2_field": label,
                    "extracted_value": val0,
                    "source_document": _src_name(c0.source_file),
                    "source_page": str(c0.source_page),
                    "sheet_category": cat,
                    "match_status": "matched",
                    "notes": ml_note,
                }
            )
        else:
            for c in clist:
                b2_rows.append(
                    {
                        "b2_field": label,
                        "extracted_value": _norm_value_for_field(field, (c.value or "").strip()),
                        "source_document": _src_name(c.source_file),
                        "source_page": str(c.source_page),
                        "sheet_category": cat,
                        "match_status": "conflict",
                        "notes": "multiple_distinct_values" + ("; " + ml_note if ml_note else ""),
                    }
                )

    ws6 = wb.create_sheet("B2_FIELD_MAP")
    _write_sheet(
        ws6,
        [
            "b2_field",
            "extracted_value",
            "source_document",
            "source_page",
            "sheet_category",
            "match_status",
            "notes",
        ],
        b2_rows,
    )

    preview_rows: list[dict[str, Any]] = []
    for field in sorted(DIRECT_VALUE_FIELDS):
        label = _preview_label(field)
        section = B2_SECTION_CODE.get(field, "")
        clist = by_field.get(field, [])
        distinct_vals = {c.value.strip() for c in clist if c.value and c.value.strip()}
        if len(distinct_vals) == 1:
            val = _norm_value_for_field(field, next(iter(distinct_vals)))
            c0 = clist[0]
            preview_rows.append(
                {
                    "b2_section": section,
                    "b2_field": label,
                    "value_to_fill": val,
                    "source_document": _src_name(c0.source_file),
                    "source_page": str(c0.source_page),
                }
            )
        elif len(distinct_vals) > 1:
            preview_rows.append(
                {
                    "b2_section": section,
                    "b2_field": label,
                    "value_to_fill": "CONFLICT — see B2_FIELD_MAP",
                    "source_document": _src_name(source_name),
                    "source_page": "",
                }
            )
        else:
            val = _norm_value_for_field(field, str(merged_extracted.get(field, "") or ""))
            preview_rows.append(
                {
                    "b2_section": section,
                    "b2_field": label,
                    "value_to_fill": val,
                    "source_document": _src_name(source_name) if val else "",
                    "source_page": "",
                }
            )

    ws7 = wb.create_sheet("B2_PREVIEW")
    _write_sheet(
        ws7,
        ["b2_section", "b2_field", "value_to_fill", "source_document", "source_page"],
        preview_rows,
    )

    missing_rows: list[dict[str, Any]] = []
    for field in sorted(DIRECT_VALUE_FIELDS):
        label = _preview_label(field)
        if field not in by_field or not by_field[field]:
            if not merged_extracted.get(field):
                missing_rows.append(
                    {
                        "expected_category": sheet_category_for_field(field),
                        "missing_or_unresolved_item": f"Missing {label}",
                        "source_document": source_name,
                        "source_page": "",
                        "reason": "no_extracted_candidate",
                    }
                )

    ws8 = wb.create_sheet("MISSING_OR_UNRESOLVED")
    _write_sheet(
        ws8,
        [
            "expected_category",
            "missing_or_unresolved_item",
            "source_document",
            "source_page",
            "reason",
        ],
        missing_rows,
    )

    wb.properties.title = f"B-2 validation {form_code} {source_name}"
    wb.properties.subject = f"generated {datetime.now(timezone.utc).isoformat()} ml={ml_info.get('ml_engine', '')}"

    for name in wb.sheetnames:
        _autosize(wb[name])

    wb.save(str(output_path))
    return output_path
